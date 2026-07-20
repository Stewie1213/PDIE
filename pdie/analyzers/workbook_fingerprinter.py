"""Fingerprinting engine for workbooks."""

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from pdie.core.fingerprint import WorkbookFingerprint, WorksheetFingerprint
from pdie.core.workbook import Workbook
from pdie.core.worksheet import Worksheet


class WorkbookFingerprinter:
    """Generates fingerprints for workbooks and worksheets."""

    @staticmethod
    def fingerprint_workbook(workbook: Workbook) -> WorkbookFingerprint:
        """Generate a fingerprint for a workbook.

        Args:
            workbook: The workbook to fingerprint.

        Returns:
            WorkbookFingerprint: The generated fingerprint.
        """
        if not workbook.file_path:
            raise ValueError("Workbook must have file_path set")

        # Load the original openpyxl workbook for advanced properties
        openpyxl_wb = load_workbook(workbook.file_path)

        # Detect macros (VBA code)
        contains_macros = workbook.file_path.suffix.lower() in (
            ".xlsm",
            ".xltm",
        )

        # Count dropdowns and formulas
        dropdown_count = 0
        formula_count = workbook.total_formulas()

        # Count merged ranges from openpyxl (more reliable)
        merged_range_count = 0
        for worksheet in workbook.worksheets.values():
            if worksheet.name in openpyxl_wb.sheetnames:
                openpyxl_ws = openpyxl_wb[worksheet.name]
                merged_range_count += len(openpyxl_ws.merged_cells.ranges)

                # Count data validations (dropdowns)
                if openpyxl_ws.data_validations:
                    dropdown_count += len(openpyxl_ws.data_validations.dataValidation)

        # Count named ranges
        named_range_count = len(openpyxl_wb.defined_names)

        # Check for images
        contains_images = False
        contains_tables = False
        for worksheet in workbook.worksheets.values():
            if worksheet.name in openpyxl_wb.sheetnames:
                openpyxl_ws = openpyxl_wb[worksheet.name]
                if openpyxl_ws._images:
                    contains_images = True
                if openpyxl_ws._tables:
                    contains_tables = True

        # Check if workbook is protected
        protected = bool(getattr(openpyxl_wb.security, "lockStructure", False))

        fingerprint = WorkbookFingerprint(
            workbook_name=workbook.name,
            file_type=workbook.file_path.suffix.lower(),
            worksheet_count=workbook.worksheet_count(),
            contains_macros=contains_macros,
            contains_images=contains_images,
            contains_tables=contains_tables,
            dropdown_count=dropdown_count,
            formula_count=formula_count,
            merged_range_count=merged_range_count,
            named_range_count=named_range_count,
            protected=protected,
        )

        return fingerprint

    @staticmethod
    def fingerprint_worksheet(
        worksheet: Worksheet, openpyxl_ws: OpenpyxlWorksheet
    ) -> WorksheetFingerprint:
        """Generate a fingerprint for a worksheet.

        Args:
            worksheet: The worksheet to fingerprint.
            openpyxl_ws: The openpyxl worksheet object for advanced properties.

        Returns:
            WorksheetFingerprint: The generated fingerprint.
        """
        # Get dimensions
        if openpyxl_ws.dimensions:
            dimensions = openpyxl_ws.dimensions.split(":")
            if len(dimensions) == 2:
                # Parse end cell to get max row and column
                end_cell = dimensions[1]
                # Extract row number (digits at the end)
                import re

                row_match = re.search(r"\d+$", end_cell)
                col_match = re.search(r"^[A-Z]+", end_cell)

                max_row = int(row_match.group()) if row_match else 0
                max_col = 0
                if col_match:
                    col_letters = col_match.group()
                    max_col = WorkbookFingerprinter._col_letters_to_number(col_letters)
            else:
                max_row = 0
                max_col = 0
        else:
            max_row = 0
            max_col = 0

        # Count merged cells
        merged_cells = len(openpyxl_ws.merged_cells.ranges)

        # Count formula cells
        formula_cells = worksheet.formula_count()

        # Count tables
        tables = len(openpyxl_ws._tables) if openpyxl_ws._tables else 0

        # Count images
        images = len(openpyxl_ws._images) if openpyxl_ws._images else 0

        # Count colors used
        color_counts = WorkbookFingerprinter._count_colors(worksheet)

        # Count data validations (dropdowns) as named regions for now
        named_regions = (
            len(openpyxl_ws.data_validations.dataValidation) if openpyxl_ws.data_validations else 0
        )

        # Count editable cells
        editable_cells = worksheet.editable_count()

        fingerprint = WorksheetFingerprint(
            name=worksheet.name,
            rows=max_row,
            columns=max_col,
            merged_cells=merged_cells,
            formula_cells=formula_cells,
            tables=tables,
            images=images,
            hidden=worksheet.hidden,
            color_counts=color_counts,
            named_regions=named_regions,
            editable_cells=editable_cells,
        )

        return fingerprint

    @staticmethod
    def _col_letters_to_number(letters: str) -> int:
        """Convert column letters to column number.

        Args:
            letters: Column letters (e.g., 'A', 'Z', 'AA', 'ZZ').

        Returns:
            int: The column number (1-based).
        """
        result = 0
        for char in letters:
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result

    @staticmethod
    def _count_colors(worksheet: Worksheet) -> dict[str, int]:
        """Count occurrences of each color in a worksheet.

        Args:
            worksheet: The worksheet to analyze.

        Returns:
            dict: Color hex codes and their counts.
        """
        color_counts: dict[str, int] = {}

        for cell in worksheet.cells.values():
            if cell.fill:
                # Normalize color format
                color = str(cell.fill).upper()
                if color and color != "NONE":
                    color_counts[color] = color_counts.get(color, 0) + 1

        return color_counts
