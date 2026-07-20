"""Reader for Excel workbooks using openpyxl."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from pdie.core.cell import Cell
from pdie.core.workbook import Workbook
from pdie.core.worksheet import Worksheet


class ExcelReader:
    """Reads Excel workbooks and converts them to PDIE models."""

    @staticmethod
    def read(file_path: Path) -> Workbook:
        """Read an Excel workbook from file.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Workbook: The loaded workbook model.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file is not a valid Excel file.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            openpyxl_wb = load_workbook(file_path)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}") from e

        workbook = Workbook(
            name=file_path.stem,
            file_path=file_path,
        )

        for idx, openpyxl_ws in enumerate(openpyxl_wb.sheetnames):
            worksheet = ExcelReader._read_worksheet(
                openpyxl_wb[openpyxl_ws], idx
            )
            workbook.worksheets[worksheet.name] = worksheet

        return workbook

    @staticmethod
    def _read_worksheet(
        openpyxl_ws: OpenpyxlWorksheet, index: int
    ) -> Worksheet:
        """Read a worksheet from openpyxl.

        Args:
            openpyxl_ws: The openpyxl worksheet object.
            index: The index of the worksheet.

        Returns:
            Worksheet: The converted worksheet model.
        """
        worksheet = Worksheet(
            name=openpyxl_ws.title,
            index=index,
            hidden=openpyxl_ws.sheet_state == "hidden",
            protected=openpyxl_ws.protection.sheet,
        )

        for row in openpyxl_ws.iter_rows():
            for openpyxl_cell in row:
                if openpyxl_cell.value is not None:
                    cell = ExcelReader._read_cell(openpyxl_cell)
                    worksheet.cells[cell.address] = cell

        return worksheet

    @staticmethod
    def _read_cell(openpyxl_cell) -> Cell:
        """Read a cell from openpyxl.

        Args:
            openpyxl_cell: The openpyxl cell object.

        Returns:
            Cell: The converted cell model.
        """
        cell = Cell(
            address=openpyxl_cell.coordinate,
            row=openpyxl_cell.row,
            column=openpyxl_cell.column,
            value=openpyxl_cell.value,
            formula=openpyxl_cell.value
            if isinstance(openpyxl_cell.value, str)
            and openpyxl_cell.value.startswith("=")
            else None,
            data_type=openpyxl_cell.data_type,
            number_format=openpyxl_cell.number_format,
            locked=openpyxl_cell.protection.locked
            if openpyxl_cell.protection
            else True,
            hidden=openpyxl_cell.protection.hidden
            if openpyxl_cell.protection
            else False,
        )

        if openpyxl_cell.font:
            cell.font_name = openpyxl_cell.font.name
            cell.font_size = openpyxl_cell.font.size
            cell.bold = openpyxl_cell.font.bold or False
            cell.italic = openpyxl_cell.font.italic or False

        if openpyxl_cell.fill:
            cell.fill = openpyxl_cell.fill.start_color.rgb if openpyxl_cell.fill.start_color else None

        if openpyxl_cell.hyperlink:
            cell.hyperlink = openpyxl_cell.hyperlink.target

        if openpyxl_cell.comment:
            cell.comment = openpyxl_cell.comment.text

        return cell
