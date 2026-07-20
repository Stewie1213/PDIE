"""Pytest configuration and fixtures."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook as OpenpyxlWorkbook


@pytest.fixture
def temp_dir() -> Path:
    """Create a temporary directory for tests."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_workbook(temp_dir: Path) -> Path:
    """Create a sample Excel workbook for testing."""
    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some sample data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Test"
    ws["B2"] = 123
    ws["B3"] = "=B2*2"

    workbook_path = temp_dir / "sample.xlsx"
    wb.save(workbook_path)
    return workbook_path
