"""Tests for the workbook fingerprinter."""

from pathlib import Path

from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.workbook.defined_name import DefinedName

from pdie.analyzers.workbook_fingerprinter import WorkbookFingerprinter
from pdie.readers.excel_reader import ExcelReader


class TestWorkbookFingerprinter:
    """Tests for WorkbookFingerprinter."""

    def test_fingerprint_basic_workbook(self, sample_workbook: Path) -> None:
        """Test fingerprinting a basic workbook."""
        workbook = ExcelReader.read(sample_workbook)
        fingerprint = WorkbookFingerprinter.fingerprint_workbook(workbook)

        assert fingerprint.workbook_name == "sample"
        assert fingerprint.file_type == ".xlsx"
        assert fingerprint.worksheet_count == 1
        assert fingerprint.contains_macros is False
        assert fingerprint.formula_count >= 0

    def test_fingerprint_workbook_properties(self, temp_dir: Path) -> None:
        """Test fingerprinting detects workbook properties."""
        # Create a workbook with merged cells and named ranges
        wb = OpenpyxlWorkbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "Data"
        ws.merge_cells("A1:B1")

        ws["A2"] = 100
        ws["B2"] = "=A2*2"

        # Add a named range
        wb.defined_names.add(DefinedName("TestRange", attr_text="Sheet1!$A$2"))

        workbook_path = temp_dir / "merged.xlsx"
        wb.save(workbook_path)

        workbook = ExcelReader.read(workbook_path)
        fingerprint = WorkbookFingerprinter.fingerprint_workbook(workbook)

        assert fingerprint.merged_range_count >= 1
        assert fingerprint.formula_count >= 1
        assert fingerprint.named_range_count >= 0

    def test_fingerprint_worksheet(self, sample_workbook: Path) -> None:
        """Test fingerprinting a worksheet."""
        from openpyxl import load_workbook

        workbook = ExcelReader.read(sample_workbook)
        openpyxl_wb = load_workbook(sample_workbook)
        openpyxl_ws = openpyxl_wb["Sheet1"]

        worksheet = workbook.get_worksheet("Sheet1")
        assert worksheet is not None

        fingerprint = WorkbookFingerprinter.fingerprint_worksheet(worksheet, openpyxl_ws)

        assert fingerprint.name == "Sheet1"
        assert fingerprint.rows > 0
        assert fingerprint.formula_cells >= 0
        assert fingerprint.hidden is False

    def test_col_letters_to_number(self) -> None:
        """Test column letter to number conversion."""
        assert WorkbookFingerprinter._col_letters_to_number("A") == 1
        assert WorkbookFingerprinter._col_letters_to_number("Z") == 26
        assert WorkbookFingerprinter._col_letters_to_number("AA") == 27
        assert WorkbookFingerprinter._col_letters_to_number("AZ") == 52
        assert WorkbookFingerprinter._col_letters_to_number("BA") == 53
        assert WorkbookFingerprinter._col_letters_to_number("ZZ") == 702

    def test_count_colors(self, temp_dir: Path) -> None:
        """Test color counting in worksheet."""
        from pdie.core.cell import Cell
        from pdie.core.worksheet import Worksheet

        # Create worksheet with colored cells
        worksheet = Worksheet(name="Sheet1", index=0)

        cell1 = Cell(address="A1", row=1, column=1, value="Red", fill="FF0000")
        cell2 = Cell(address="A2", row=2, column=1, value="Red", fill="FF0000")
        cell3 = Cell(address="A3", row=3, column=1, value="Blue", fill="0000FF")

        worksheet.cells[cell1.address] = cell1
        worksheet.cells[cell2.address] = cell2
        worksheet.cells[cell3.address] = cell3

        color_counts = WorkbookFingerprinter._count_colors(worksheet)

        assert "FF0000" in color_counts
        assert color_counts["FF0000"] == 2
        assert "0000FF" in color_counts
        assert color_counts["0000FF"] == 1

    def test_fingerprint_no_file_path_raises(self) -> None:
        """Test that fingerprinting without file_path raises."""
        from pdie.core.workbook import Workbook

        workbook = Workbook(name="test", file_path=None)

        try:
            WorkbookFingerprinter.fingerprint_workbook(workbook)
            raise AssertionError("Should have raised ValueError")
        except ValueError as e:
            assert "file_path" in str(e)
